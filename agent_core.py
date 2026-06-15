"""Core logic for the Precursive data-entry agent.

Refactored from the standalone script so a UI (Streamlit) can drive it:
- credentials, the Excel path, and the column selection are passed in as arguments
- progress is reported through an optional callback instead of printing
- no module-level file side effects (each run owns its own log files)

Behaviour is otherwise identical to the original: same task accounting,
reconcile(), fail_all(), resume-via-progress-file, and incremental CSV writing.
"""

from browser_use import Agent, Browser, BrowserProfile, ChatAnthropic
from dotenv import load_dotenv
from pydantic import BaseModel
from typing import Literal, Callable, Optional
import openpyxl
import csv
import os
from datetime import datetime
from collections import defaultdict

load_dotenv()  # loads ANTHROPIC_API_KEY (and anything else) from a local .env

MAX_STEPS_PER_ENTRY = 25


# ---------------- data models ----------------
class TaskStatus(BaseModel):
    id: int
    tag: str
    hours: float
    status: Literal["logged", "skipped", "failed"]
    reason: str = ""


class EntryResult(BaseModel):
    tasks: list[TaskStatus]


class TaskOutcome(BaseModel):
    tag: str
    hours: float
    date: str
    day: str
    project: str
    status: str     # "skipped" or "failed"
    reason: str


SPEED_OPTIMIZATION_PROMPT = """
Speed optimization instructions:
- Be extremely concise and direct in your responses.
- Get to the goal as quickly as possible.
- Assume pages are ready as soon as key interactive elements are available.
- Use multi-action sequences whenever possible to reduce steps.

Chain-of-Draft reasoning (keep internal reasoning minimal):
- In thinking, evaluation_previous_goal, and next_goal, write terse drafts, NOT sentences.
- Limit each reasoning point to ~5 words. Notes, not prose. Drop articles and filler.
- Record only the decisive signal: what changed, what's next. Do not restate page contents.
- Example next_goal: "Click login; enter email" — NOT a full sentence.

- Once you are done with the entire process just say "I'm Done Yazan". No summary, nothing else.
"""

TASK_TEMPLATE = """
Follow these steps to complete the process:
1) Go to https://appliedai.lightning.force.com/lightning/n/preempt__My_Precursive. If you are NOT already logged in, log in with the details provided then click submit. If you are asked for an authentication code then wait just 5 seconds as it will be written automatically. If already logged in, skip login entirely.
2) In the middle of the page you will see a Date Navigation Control in the format DD/MM/YYYY. Change the date to {date} (also DD/MM/YYYY). If you are already on that date, ignore this step.
3) In the timesheet data entry grid, projects are rows on the left with their tasks; dates are columns. Find project {project}, locate {day} ({date}), and enter the hours from the task list below (usually shown as "0h"). Press Enter after each entry. Scroll if not all tasks are visible.
4) Tasks don't need to match exactly. If logically similar, enter them. If not, skip and move on.

Task list:
{task_hours}

REPORTING (required): report the status of EVERY task above, one entry per task, using the task numbers exactly as listed. For each task return:
- id: the task number (1, 2, 3, ...)
- tag: the task name as given
- hours: the hours value as given
- status: "logged" if you successfully entered the hours, "skipped" if no matching row existed, "failed" if you tried but something went wrong
- reason: brief explanation for skipped/failed (leave empty for logged)
Do not omit any task. If you are unsure whether a task was logged, mark it "failed".
"""


# ---------------- excel parsing ----------------
def load_excel_entries(path, columns=None):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    DATE_START_COL = 4
    days = {}
    dates = {}

    for col in ws.iter_cols(min_row=2, max_row=3, min_col=DATE_START_COL):
        col_idx = col[0].column
        day_val = col[0].value
        date_val = col[1].value
        if day_val is None and date_val is None:
            continue
        day_str = str(day_val).strip() if day_val else ""
        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%d/%m/%Y")
        elif date_val:
            raw = str(date_val).strip()
            date_str = raw + f"/{datetime.now().year}" if raw.count("/") == 1 else raw
        else:
            date_str = ""
        days[col_idx] = day_str
        dates[col_idx] = date_str

    date_cols = sorted(dates.keys())  # position 1 = first date column

    if columns is None:
        positions = range(1, len(date_cols) + 1)   # all
    elif isinstance(columns, int):
        positions = range(1, columns + 1)          # first N columns
    else:
        positions = columns                        # explicit list of positions
    selected_cols = [date_cols[p - 1] for p in positions if 1 <= p <= len(date_cols)]

    date_project_tasks = defaultdict(lambda: defaultdict(list))
    date_project_client = {}

    for row in ws.iter_rows(min_row=4, values_only=True):
        client, project, tag = row[0], row[1], row[2]
        if project is None:
            continue
        for col_idx in date_cols:
            hours_val = row[col_idx - 1]
            if hours_val is None or hours_val == "" or hours_val == 0:
                continue
            try:
                hours = float(hours_val)
            except (ValueError, TypeError):
                continue
            if hours > 0:
                date_project_tasks[col_idx][str(project).strip()].append(
                    (str(tag).strip() if tag else "General", hours)
                )
                date_project_client[(col_idx, str(project).strip())] = (
                    str(client).strip() if client else ""
                )

    entries = []
    for col_idx in selected_cols:
        for project, task_list in date_project_tasks[col_idx].items():
            total = sum(h for _, h in task_list)
            tasks = [
                {"id": i + 1, "tag": tag, "hours": hours}
                for i, (tag, hours) in enumerate(task_list)
            ]
            task_lines = "\n".join(f"{t['id']}) {t['tag']}: {t['hours']}h" for t in tasks)
            entries.append({
                "date": dates[col_idx],
                "day": days[col_idx],
                "project": project,
                "client": date_project_client.get((col_idx, project), ""),
                "task_hours": task_lines,
                "tasks": tasks,
                "total_hours": total,
            })

    return entries


# ---------------- helpers ----------------
def entry_key(entry):
    return f"{entry['date']}|{entry['project']}"


def make_outcome(entry, item, status, reason):
    return TaskOutcome(
        tag=item["tag"], hours=item["hours"],
        date=entry["date"], day=entry["day"], project=entry["project"],
        status=status, reason=reason or "",
    )


def fail_all(entry, reason, status="failed"):
    return [make_outcome(entry, it, status, reason) for it in entry["tasks"]]


def reconcile(entry, reported_tasks):
    by_id = {t.id: t for t in (reported_tasks or [])}
    failures = []
    for it in entry["tasks"]:
        rep = by_id.get(it["id"])
        if rep is None:
            failures.append(make_outcome(entry, it, "failed", "not reported by agent"))
        elif rep.status != "logged":
            status = rep.status if rep.status in ("skipped", "failed") else "failed"
            failures.append(make_outcome(entry, it, status, rep.reason or "no reason given"))
    return failures


def load_progress(progress_file):
    if not os.path.exists(progress_file):
        return set()
    with open(progress_file, encoding="utf-8") as f:
        return {line.strip() for line in f if line.strip()}


def mark_done(progress_file, key):
    with open(progress_file, "a", encoding="utf-8") as f:
        f.write(key + "\n")


# ---------------- the runner ----------------
async def run_entries(
    excel_path: str,
    columns,                       # None | int | list[int]
    email: str,
    password: str,
    *,
    headless: bool = False,
    logs_dir: str = "logs",
    progress_cb: Optional[Callable] = None,
):
    """Process the selected columns. Calls progress_cb after each entry with:
       (done_count, total, failures_list_of_dicts, total_tokens, total_cost, label).
    Returns a summary dict including the CSV path and the full failures list."""
    os.makedirs(logs_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    fail_csv = os.path.join(logs_dir, f"missed_{ts}.csv")
    progress_file = os.path.join(logs_dir, f"progress_{os.path.basename(excel_path)}.txt")

    entries = load_excel_entries(excel_path, columns=columns)
    done = load_progress(progress_file)
    total = len(entries)

    fail_fp = open(fail_csv, "a", newline="", encoding="utf-8")
    writer = csv.writer(fail_fp)
    writer.writerow(["date", "day", "project", "tag", "hours", "status", "reason"])
    fail_fp.flush()

    claude_llm = ChatAnthropic(model="claude-sonnet-4-6")
    page_extraction_model = ChatAnthropic(model="claude-haiku-4-5-20251001")

    browser = Browser(
        keep_alive=True,  # log in once, reuse the session for every entry
        browser_profile=BrowserProfile(
            minimum_wait_page_load_time=0.1,
            wait_between_actions=0.1,
            headless=headless,
            is_local=False
        ),
        allowed_domains=["*.salesforce.com", "*.lightning.force.com"],
    )
    await browser.start()

    all_failures = []      # list of dicts, for the UI table + return value
    total_tokens = 0
    total_cost = 0.0

    def emit(done_count, label):
        if progress_cb:
            progress_cb(done_count, total, list(all_failures), total_tokens, total_cost, label)

    try:
        for i, entry in enumerate(entries):
            key = entry_key(entry)
            label = f"{entry['day']} {entry['date']} | {entry['project']}"

            if key in done:
                emit(i + 1, f"Skipped (already done): {label}")
                continue

            task = TASK_TEMPLATE.format(
                date=entry["date"], day=entry["day"],
                project=entry["project"], task_hours=entry["task_hours"],
            )

            agent = Agent(
                task=task,
                sensitive_data={"x_user": email, "x_pass": password},
                browser=browser,
                llm=claude_llm,
                use_vision=True,
                vision_detail_level="auto",
                page_extraction_llm=page_extraction_model,
                use_thinking=False,
                flash_mode=True,
                extend_system_message=SPEED_OPTIMIZATION_PROMPT,
                max_history_items=None,
                directly_open_url=True,
                calculate_cost=True,
                generate_gif=False,
                max_failures=3,
                max_actions_per_step=5,
                output_model_schema=EntryResult,
            )

            try:
                run = await agent.run(max_steps=MAX_STEPS_PER_ENTRY)
                result = run.structured_output
                if result is None:
                    entry_failures = fail_all(entry, "agent returned no structured output")
                else:
                    entry_failures = reconcile(entry, result.tasks)
            except Exception as e:
                entry_failures = fail_all(entry, f"agent crashed: {e}")

            try:
                s = await agent.token_cost_service.get_usage_summary()
                total_tokens += s.total_tokens
                total_cost += s.total_cost
            except Exception:
                pass

            for t in entry_failures:
                writer.writerow([t.date, t.day, t.project, t.tag, t.hours, t.status, t.reason])
                fail_fp.flush()
                all_failures.append(t.model_dump())

            if not entry_failures:
                mark_done(progress_file, key)

            emit(i + 1, label)
    finally:
        fail_fp.close()
        await browser.kill()

    return {
        "csv_path": fail_csv,
        "failures": all_failures,
        "total": total,
        "tokens": total_tokens,
        "cost": total_cost,
    }