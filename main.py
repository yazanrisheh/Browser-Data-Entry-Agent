from browser_use import Agent, Browser, BrowserProfile, ChatAnthropic, ChatOpenAI
from dotenv import load_dotenv
import asyncio
import openpyxl
import time
import os
import sys
from datetime import datetime

load_dotenv()

LOG_FILE = "run_log_all.txt"

class Tee:
    def __init__(self, file):
        self.file = file
        self.stdout = sys.__stdout__

    def write(self, data):
        self.stdout.write(data)
        self.file.write(data)

    def flush(self):
        self.stdout.flush()
        self.file.flush()

# EXCEL_FILE = "Timely Data - Yazan - Timely Data - Yazan.xlsx"
EXCEL_FILE = "Yazan data.xlsx"
ROWS = None  # Set to None to process all rows (one per date column with hours)

SPEED_OPTIMIZATION_PROMPT = """
Speed optimization instructions:
- Be extremely concise and direct in your responses
- Get to the goal as quickly as possible.
- Assume pages are ready as soon as key interactive elements are available
- Use multi-action sequences whenever possible to reduce steps
- Once you are done with entire process just say "I'm Done Yazan". No need to say anything else or any give any summary
"""

def load_excel_entries(path, max_entries=None):
    """
    Returns a list of dicts, one per (date_column, project) combination that has hours.
    Each dict has: date, day, project, client, task_hours (formatted string), total_hours.

    Excel layout:
      Row 2: day names  (Mon, Tue, ...)  starting from col D (index 3)
      Row 3: dates      (06/04, 07/04, ...) starting from col D
      Row 4+: data rows with Client (col A), Project (col B), Hour Tags (col C), then hours per date column
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Read day and date headers from row 2 and row 3 (1-indexed in openpyxl)
    # Data columns start at column index 4 (col D = index 4 in openpyxl, 1-based)
    DATE_START_COL = 4  # col D

    days = {}   # col_index -> day string e.g. "Monday"
    dates = {}  # col_index -> date string e.g. "06/04/2026"

    for col in ws.iter_cols(min_row=2, max_row=3, min_col=DATE_START_COL):
        col_idx = col[0].column
        day_val = col[0].value   # row 2
        date_val = col[1].value  # row 3

        if day_val is None and date_val is None:
            continue

        # Normalise day
        day_str = str(day_val).strip() if day_val else ""

        # Normalise date — could be a datetime object or a string like "06/04"
        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%d/%m/%Y")
        elif date_val:
            raw = str(date_val).strip()
            # If only DD/MM append current year
            if raw.count("/") == 1:
                date_str = raw + f"/{datetime.now().year}"
            else:
                date_str = raw
        else:
            date_str = ""

        days[col_idx] = day_str
        dates[col_idx] = date_str

    date_cols = sorted(dates.keys())

    # Read data rows starting from row 4
    # Group by date: for each date column collect all (project, tag, hours) entries
    # Then build one task entry per (date, project) pair
    from collections import defaultdict
    date_project_tasks = defaultdict(lambda: defaultdict(list))
    # date_project_tasks[col_idx][project] = [(tag, hours), ...]
    date_project_client = {}  # (col_idx, project) -> client

    for row in ws.iter_rows(min_row=4, values_only=True):
        client = row[0]   # col A
        project = row[1]  # col B
        tag = row[2]      # col C

        if project is None:
            continue

        for col_idx in date_cols:
            hours_val = row[col_idx - 1]  # openpyxl row is 0-indexed tuple
            if hours_val is None or hours_val == "" or hours_val == 0:
                continue
            try:
                hours = float(hours_val)
            except (ValueError, TypeError):
                continue
            if hours > 0:
                date_project_tasks[col_idx][str(project).strip()].append(
                    (str(tag).strip() if tag else "General", hours)
                )
                date_project_client[(col_idx, str(project).strip())] = str(client).strip() if client else ""

    # Flatten into task entries
    entries = []
    for col_idx in date_cols:
        for project, task_list in date_project_tasks[col_idx].items():
            total = sum(h for _, h in task_list)
            # Build numbered task-hours string
            task_lines = "\n".join(
                f"{i + 1}) {tag}: {hours}h"
                for i, (tag, hours) in enumerate(task_list)
            )
            entries.append({
                "date": dates[col_idx],
                "day": days[col_idx],
                "project": project,
                "client": date_project_client.get((col_idx, project), ""),
                "task_hours": task_lines,
                "total_hours": total,
            })

    if max_entries is not None:
        entries = entries[:max_entries]

    return entries


async def main():
    log_file = open(LOG_FILE, "a", encoding="utf-8")
    sys.stdout = Tee(log_file)

    claude_llm = ChatAnthropic(model="claude-sonnet-4-6")
    openai_llm = ChatOpenAI(model="gpt-4o-mini")
    page_extraction_model = ChatAnthropic(model="claude-haiku-4-5-20251001")

    browser_profile_info = BrowserProfile(
        minimum_wait_page_load_time=0.1,
        wait_between_actions=0.1,
        headless=False,
    )

    browser = Browser(
        keep_alive=True,
        browser_profile=browser_profile_info,
        allowed_domains=["*.salesforce.com", "*.lightning.force.com"],
    )
    await browser.start()

    entries = load_excel_entries(EXCEL_FILE, max_entries=ROWS)
    print(f"Loaded {len(entries)} entries from {EXCEL_FILE}")

    overall_start = time.time()
    total_tokens_all = 0
    total_cost_all = 0.0

    for i, entry in enumerate(entries):
        task = """
Follow these steps to complete the process:
1) Go to https://appliedai.lightning.force.com/lightning/n/preempt__My_Precursive. If you are NOT already logged in, log in with the details provided then click submit. If you are asked for authentication code then wait just 5 seconds as it will be written automatically. If already logged in, skip login entirely.
2) In the middle of the page, you will see Date Navigation Control in the format format DD/MM/YYYY. Change the date to {date} also in format DD/MM/YYYY. If you are already on the same date then ignore this step.
3) In the timesheet data entry grid, the project names are on the left side along with their tasks as rows whereas the dates are columns. Look for the project {project} and the correct day & date for {day} & {date} then write the correct hours for each task {task_hours} which are most of the time "0h". Press enter after writing the hour for each task
4) Tasks do not need to have exact match. If they are logically similar then u can do the data entry for it otherwise ignore. Do not overthink and waste time on thinking twice. If its simply not logically same then do the next one.
""".format(
            date=entry["date"],
            day=entry["day"],
            project=entry["project"],
            task_hours=entry["task_hours"],
        )

        print(f"\n--- Entry {i + 1}/{len(entries)}: {entry['day']} {entry['date']} | {entry['project']} ({entry['total_hours']}h) ---")
        print(f"Tasks:\n{entry['task_hours']}")
        row_start = time.time()

        agent = Agent(
            task=task,
            sensitive_data={
                "x_user": os.getenv("EMAIL"),
                "x_pass": os.getenv("PASS"),
            },
            browser=browser,
            llm=claude_llm,
            use_vision=True,
            vision_detail_level="auto",
            page_extraction_llm=page_extraction_model,
            use_thinking=False,
            flash_mode=True,
            extend_system_message=SPEED_OPTIMIZATION_PROMPT,
            max_history_items=None,
            directly_open_url=True,
            calculate_cost=True,
            generate_gif=False,
            max_failures=5,
            max_actions_per_step=5,
        )

        run = await agent.run()

        print("Output: ", run.final_result())
        row_elapsed = time.time() - row_start
        s = await agent.token_cost_service.get_usage_summary()
        total_tokens_all += s.total_tokens
        total_cost_all += s.total_cost
        print(f"Entry {i + 1} completed in {row_elapsed:.1f}s | tokens: {s.total_tokens:,} | cost: ${s.total_cost:.4f}")

    overall_elapsed = time.time() - overall_start
    print(f"\n--- Overall ---")
    print(f"Entries processed : {len(entries)}")
    print(f"Total time        : {overall_elapsed:.1f}s")
    print(f"Total tokens      : {total_tokens_all:,}")
    print(f"Total cost        : ${total_cost_all:.4f}")

    await browser.kill()
    sys.stdout = sys.__stdout__
    log_file.close()

if __name__ == "__main__":
    asyncio.run(main())
